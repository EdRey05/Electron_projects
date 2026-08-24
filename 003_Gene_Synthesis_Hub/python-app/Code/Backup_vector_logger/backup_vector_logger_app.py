"""
Bio Basic Inc. - Canada
Made by: Eduardo Reyes, Ph.D.
Contact: ed5reyes@outlook.com

Version: 1.6
Date: Mar 08, 2026
Notes: Same functionality as v1.5 but adapted to new app hub.
"""

import pandas as pd
import configparser
import streamlit as st
import os
import re
import shutil
import sys
import warnings
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill

# Suppress the specific openpyxl warning about conditional formatting extensions.
warnings.filterwarnings(
    "ignore", category=UserWarning, module="openpyxl.worksheet._reader"
)

# App setup and layout
st.set_page_config(page_title="Backup Logger", layout="wide")
st.title("WGK/BATCH Gene Backup Log System")
st.markdown(
    """<hr style="border: none; height: 5px; background-color: #444;">""",
    unsafe_allow_html=True,
)

# Define the subfolder where temporary files are stored
TEMP_SUBFOLDER = "temp_logs"

# Define network and local file paths
LOG_FILENAMES = {
    "WGK": "WGK Gene Backup Log.xlsx",
    "BATCH": "BATCH Gene Backup Log.xlsx",
}

# Row 1 cols and divider
row1_col1, row1_col2, row1_col3, row1_col4 = st.columns(
    4,
    vertical_alignment="bottom",
    gap="large",
)
st.markdown(
    """<hr style="border: none; height: 5px; background-color: #444;">""",
    unsafe_allow_html=True,
)

# Row 2 cols (no divider)
row2_col1, row2_col2, row2_col3, row2_col4, row2_col5 = st.columns(
    [1, 3.1, 1, 8, 1],
    gap="small",
)

# Row 3 cols and divider
row3_col1, row3_col2, row3_col3, row3_col4, row3_col5 = st.columns(
    [1, 4, 4, 2, 1],
    gap="small",
)
st.markdown(
    """<hr style="border: none; height: 5px; background-color: #444;">""",
    unsafe_allow_html=True,
)

# Row 4 container
row4 = st.container()


def parse_app_args():
    """
    Parses command line arguments passed by the Hub.
    Returns the config file path and a dictionary of drive mappings.
    """
    config_file = "default_directories.ini"  # Default fallback
    drive_map = {}

    # sys.argv[0] is the script name
    for arg in sys.argv[1:]:
        if arg.startswith("--config="):
            config_file = arg.split("=", 1)[1]
        elif "=" in arg and not arg.startswith("--"):
            key, val = arg.split("=", 1)
            drive_map[key] = val

    # Fallback: if config path is relative, try to find it in parent directories
    if not os.path.isabs(config_file) and not os.path.exists(config_file):
        # Assuming standard structure: Hub/Code/App -> Hub/default_directories.ini
        parent_config = os.path.abspath(os.path.join("..", "..", config_file))
        if os.path.exists(parent_config):
            config_file = parent_config

    return config_file, drive_map


def get_path_from_config(config_key, drive_letter):
    """
    Reads 'default_directories.ini' to find the relative path for a given key,
    then prepends the provided drive letter.
    Expected format in file: Key = Label, "Path"
    """
    config_path, _ = parse_app_args()
    if not os.path.exists(config_path):
        st.error(f"Configuration file not found: {config_path}")
        st.stop()

    config = configparser.ConfigParser()
    config.optionxform = str  # Preserve case sensitivity
    config.read(config_path, encoding="utf-8")

    if "Paths" in config and config_key in config["Paths"]:
        raw_value = config["Paths"][config_key]
        # Parse: Label, "Path"
        if "," in raw_value:
            parts = [p.strip() for p in raw_value.split(",", 1)]
            if len(parts) == 2:
                # We ignore the label here and trust the drive_letter passed to the app
                # corresponds to the label intended for these files (B8. Gene Synthesis)
                subpath = parts[1].strip('"').strip("'")
                return os.path.join(drive_letter, subpath)
    return None


@st.cache_resource(show_spinner="Syncing log files and database with network drive...")
def sync_log_files_on_startup():
    """
    Copies log files and the invoicing database from the network drive to the local temp folder on startup.
    Creates a single backup of each original log file, overwriting any previous backup.
    """
    if "drive_letter" not in st.session_state:
        _, drive_map = parse_app_args()
        # Look for the specific label used by this app
        target_label = "B8. Gene Synthesis"

        if target_label in drive_map:
            st.session_state.drive_letter = drive_map[target_label]
        else:
            st.error(
                "Network drive letter not provided. Please launch this app via the App Hub."
            )
            st.stop()

    drive = st.session_state.drive_letter

    # Resolve Network Log Directory from Config
    network_log_dir = get_path_from_config("Vector_Storage_Dir", drive)
    if not network_log_dir:
        st.error("Could not resolve 'Vector_Storage_Dir' from config.")
        st.stop()

    if not os.path.isdir(network_log_dir):
        st.error(f"Network directory not found: {network_log_dir}")
        st.stop()

    # Ensure the local temporary directory exists
    if not os.path.exists(TEMP_SUBFOLDER):
        os.makedirs(TEMP_SUBFOLDER)

    # --- NEW: Copy Gene Invoicing Database from Network ---
    # This replaces the logic previously handled by the .bat file
    invoicing_db_source = get_path_from_config("Gene_Invoicing_Db", drive)
    if not invoicing_db_source:
        st.error("Could not resolve 'Gene_Invoicing_Db' from config.")

    invoicing_db_local = os.path.join(TEMP_SUBFOLDER, "Gene Invoicing database.xlsx")

    if os.path.exists(invoicing_db_source):
        try:
            shutil.copy2(invoicing_db_source, invoicing_db_local)
        except Exception as e:
            st.error(f"Failed to copy invoicing database: {e}")
    else:
        st.warning(f"Invoicing database not found at: {invoicing_db_source}")
    # -------------------------------------------------------

    for log_name in LOG_FILENAMES.values():
        source_path = os.path.join(network_log_dir, log_name)
        local_path = os.path.join(TEMP_SUBFOLDER, log_name)
        backup_path = os.path.join(
            TEMP_SUBFOLDER, f"{os.path.splitext(log_name)[0]}_backup.xlsx"
        )

        try:
            if os.path.exists(source_path):
                # Copy from network to local for editing
                shutil.copy2(source_path, local_path)
                # Create a timestamped backup of the original network file
                shutil.copy2(source_path, backup_path)
            elif not os.path.exists(local_path):
                # If the file doesn't exist on the network or locally, create a new blank one
                pd.DataFrame(
                    columns=[
                        "Box name",
                        "Location",
                        "Gene",
                        "Tube ID",
                        "Date",
                        "Amount",
                        "Notes",
                    ]
                ).to_excel(local_path, sheet_name="Log", index=False)
        except Exception as e:
            st.error(f"Failed to copy or back up '{log_name}': {e}")


def format_date_val(val):
    """Helper to format Excel dates into strings (YYYY-MM-DD)."""
    if pd.isna(val) or val == "":
        return ""
    if isinstance(val, datetime) or isinstance(val, pd.Timestamp):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


@st.cache_data(show_spinner=False)
def load_reference_data_from_path(ref_path: str, exceptions_path: str):
    # Desired column names
    col_names = [
        "BBI_SO",  # O
        "BBI_PO",  # P
        "Customer_No",  # Q
        "Length",  # S
        "Vector",  # T
        "Plasmid_No",  # V
        "Amount",  # W
        "Date",  # AB
    ]

    # Load all needed columns from Excel by letter range
    df_main = pd.read_excel(
        ref_path, sheet_name="Sort Out Blank", usecols="O,P,Q,S,T,V,W,AB"
    )
    df_main.columns = col_names

    # Apply date formatting immediately upon load
    df_main["Date"] = df_main["Date"].apply(format_date_val)

    # Load Exceptions CSV
    try:
        df_exceptions_raw = pd.read_csv(exceptions_path, header=0)
        df_exceptions = df_exceptions_raw.iloc[:, [0, 1, 2, 4, 5, 7, 8]].copy()
        df_exceptions["Date"] = ""
        df_exceptions.columns = col_names
    except FileNotFoundError:
        df_exceptions = pd.DataFrame(columns=col_names)

    return df_main, df_exceptions


def copy_log_to_network(local_log_path: str):
    """Copies the updated local log file back to the network drive."""
    if "drive_letter" not in st.session_state:
        st.warning("Cannot sync to network: Drive letter not available.")
        return

    drive = st.session_state.drive_letter

    # Resolve Network Log Directory from Config
    network_log_dir = get_path_from_config("Vector_Storage_Dir", drive)
    # Note: We assume network_log_dir is valid here since we checked it at startup

    log_filename = os.path.basename(local_log_path)
    network_path = os.path.join(network_log_dir, log_filename)

    try:
        shutil.copy2(local_log_path, network_path)
        st.toast(f"Synced '{log_filename}' to network.", icon="☁️")
    except Exception as e:
        st.error(f"Failed to copy '{log_filename}' to network: {e}")


def save_and_format_log(file_path: str, df: pd.DataFrame):
    """
    Saves the DataFrame to the Excel file, replacing the 'Log' sheet
    and reapplying column widths and conditional formatting.
    """
    # Write to a temporary sheet first
    with pd.ExcelWriter(
        file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name="Log_temp", index=False)

    wb = load_workbook(file_path)

    # Remove old Log sheet if it exists and rename temp sheet
    if "Log" in wb.sheetnames:
        wb.remove(wb["Log"])

    if "Log_temp" in wb.sheetnames:
        ws = wb["Log_temp"]
        ws.title = "Log"
    else:
        return

    # Autofit columns
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[col_letter].width = max_length + 2

    # Conditional Formatting for duplicates
    for col_letter, color in [("C", "FFFFFF00"), ("D", "FFFFC000")]:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        dxf = DifferentialStyle(fill=fill)
        rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=None)
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    wb.save(file_path)


# --- Main App Logic ---
sync_log_files_on_startup()

# Perform the initial data load and file deletion only once per session.
if "data_loaded" not in st.session_state:
    with st.spinner("Loading reference data..."):
        ref_path = os.path.join(TEMP_SUBFOLDER, "Gene Invoicing database.xlsx")
        exceptions_path = os.path.join(TEMP_SUBFOLDER, "Exceptions_log.csv")

        # Call the cached function to load data
        df_ref, df_exc = load_reference_data_from_path(ref_path, exceptions_path)

        # Store data in session state
        st.session_state.df_ref = df_ref
        st.session_state.df_exc = df_exc

        # Delete the file and set the flag after successful loading
        if os.path.exists(ref_path):
            os.remove(ref_path)
        st.session_state.data_loaded = True

df_ref, df_exc = st.session_state.df_ref, st.session_state.df_exc


def create_new_box(df_log, current_box_name, box_options=None):
    try:
        # Use last box option if current_box_name empty or invalid
        if not current_box_name or current_box_name.strip() == "":
            current_box_name = box_options[-1]  # last existing box name

        prefix, num = current_box_name.rsplit(" ", 1)
        next_box_num = int(num) + 1
        new_box_name = f"{prefix} {next_box_num}"
    except Exception:
        new_box_name = "Box 1"

    # Check if new_box_name already exists; if yes, increment until unique
    existing_boxes = set(df_log.iloc[:, 0].unique())
    while new_box_name in existing_boxes:
        try:
            prefix, num = new_box_name.rsplit(" ", 1)
            next_box_num = int(num) + 1
            new_box_name = f"{prefix} {next_box_num}"
        except Exception:
            # Just append "_new" if something weird happens
            new_box_name = new_box_name + "_new"
            break

    positions = [f"{row}{col}" for row in "ABCDEFGHI" for col in range(1, 10)]

    new_rows = pd.DataFrame(
        {
            df_log.columns[0]: [new_box_name] * len(positions),
            "Location": positions,
            "Gene": ["" for _ in positions],
            "Tube ID": ["" for _ in positions],
            "Date": ["" for _ in positions],
            "Amount": ["" for _ in positions],
            "Notes": ["" for _ in positions],
        }
    )

    updated_log = pd.concat([df_log, new_rows], ignore_index=True)
    return updated_log, new_box_name


def is_nonempty(val):
    if isinstance(val, str):
        return val.strip() != ""
    return not pd.isna(val)


def natural_sort_key(s):
    """
    A sort key function for natural sorting of strings containing numbers.
    e.g., sorts 'Box 10' after 'Box 2'.
    """
    return [
        int(text) if text.isdigit() else text.lower()
        for text in re.split("([0-9]+)", s)
    ]


# Initial select widget for gene type (different excel log file)
with row1_col1:
    gene_type = st.selectbox(
        "Select Gene Type (Log file):",
        options=["", "WGK", "BATCH"],
        format_func=lambda x: "Select Gene Type" if x == "" else x,
    )

# Initial select widget for box name
with row1_col2:
    box_options = [""]
    box_name = None

    if gene_type:
        log_filename = LOG_FILENAMES.get(gene_type)
        if log_filename:
            filename = os.path.join(TEMP_SUBFOLDER, log_filename)
            try:
                df_log = pd.read_excel(filename, sheet_name="Log")
                if "Date" not in df_log.columns:
                    df_log.insert(4, "Date", "")
                df_log = df_log.fillna("")
                unique_boxes = df_log.iloc[:, 0].dropna().unique().tolist()
                unique_boxes.sort(key=natural_sort_key)
                box_options += unique_boxes
            except Exception as e:
                st.error(f"Error reading file '{filename}': {e}")

    if "box_name_selected" not in st.session_state:
        st.session_state.box_name_selected = ""

    # If a new box was just created, set it as the selected one
    if st.session_state.get("box_just_created") and gene_type:
        st.session_state.box_name_selected = st.session_state.get("new_box_name")
        st.session_state["box_just_created"] = False

    box_name = st.selectbox(
        "Select Box Name:",
        options=box_options,
        index=(
            box_options.index(st.session_state.box_name_selected)
            if st.session_state.box_name_selected in box_options
            else 0
        ),
        format_func=lambda x: "Select Box Name" if x == "" else x,
        key="box_name_selected",
    )

# Create new Box Button
with row1_col3:
    if gene_type:
        if st.button("Create New Box", type="primary"):
            # Use the last (most recent) box name in the dropdown to determine the new one
            if len(box_options) > 1:
                last_existing_box = box_options[-1]
            else:
                last_existing_box = "Box 1"

            updated_log_df, new_box = create_new_box(
                df_log, last_existing_box, box_options
            )

            save_and_format_log(filename, updated_log_df)

            st.session_state["box_just_created"] = True
            st.session_state["new_box_name"] = new_box
            # Sync the newly created box to the network immediately
            copy_log_to_network(filename)
            st.rerun()

# Stop unless both gene type (WGK or BATCH) and box are selected
if not gene_type or not box_name:
    st.stop()

# Filter df for selected box
filtered_df = df_log[df_log.iloc[:, 0] == box_name].copy().reset_index(drop=True)

# Assign to session state
if (
    "box_df" not in st.session_state
    or st.session_state.get("current_box_name") != box_name
):
    st.session_state.box_df = filtered_df
    st.session_state.current_box_name = box_name
    st.session_state.selected_spot = None
current_box_df = st.session_state.box_df

# Widgets for tube info entry
with row2_col2:
    st.subheader("Enter Tube Information")
    gene_identifier = st.text_input("Tube identifier")
    id_type = st.radio(" ", ["Tube ID", "Gene"], horizontal=True)

    amount = st.selectbox("Amount", options=["1 ug", "2 ug", "4 ug", "10 ug", "Other"])
    if amount == "Other":
        custom_amount = st.text_input("Custom amount")
        amount = custom_amount if custom_amount else ""

    notes = st.text_area("Notes (optional)")

# Grid for box spot selection
with row2_col4:
    st.subheader("Select a Spot in the Box")
    rows = list("ABCDEFGHI")
    cols = list(range(1, 10))

    for row in rows:
        button_row = st.columns(9)
        for idx, col_num in enumerate(cols):
            spot_label = f"{row}{col_num}"
            spot_data = current_box_df[current_box_df["Location"] == spot_label].iloc[0]
            is_filled = is_nonempty(spot_data["Gene"]) or is_nonempty(
                spot_data["Tube ID"]
            )

            btn_type = (
                "primary"
                if st.session_state.get("selected_spot") == spot_label
                else "secondary"
            )
            btn_disabled = is_filled

            with button_row[idx]:
                if st.button(
                    spot_label, type=btn_type, disabled=btn_disabled, key=spot_label
                ):
                    st.session_state.selected_spot = spot_label
                    st.rerun()

# Query to find second identifier in shipping log
queried_gene = ""
queried_tube_id = ""
queried_date = ""

if gene_identifier.strip():
    # Determine lookup column
    lookup_col = "Plasmid_No" if id_type == "Tube ID" else "BBI_PO"
    normalized_id = gene_identifier.strip().upper()

    # Try main reference first
    result = df_ref[
        df_ref[lookup_col].astype(str).str.strip().str.upper() == normalized_id
    ]

    # If not found, try exceptions
    if result.empty and not df_exc.empty:
        result = df_exc[
            df_exc[lookup_col].astype(str).str.strip().str.upper() == normalized_id
        ]

    # Extract results if found
    if not result.empty:
        queried_gene = result.iloc[0]["BBI_PO"]
        queried_tube_id = result.iloc[0]["Plasmid_No"]
        queried_date = result.iloc[0]["Date"]
        # Date is already formatted as a clean string in load_reference_data_from_path
    else:
        queried_gene = None
        queried_tube_id = None
        queried_date = None


def styled_message(text, message_type="success"):
    """Generates HTML for a styled message box."""
    if message_type == "success":
        bg_color, border_color, text_color = "#D4EDDA", "#C3E6CB", "#155724"
    elif message_type == "error":
        bg_color, border_color, text_color = "#F8D7DA", "#F5C6CB", "#721C24"
    else:  # warning/neutral
        bg_color, border_color, text_color = "#FFF3CD", "#FFEEBA", "#856404"

    html = f"""
    <div style="
        background-color: {bg_color};
        border: 1px solid {border_color};
        color: {text_color};
        padding: 1rem;
        border-radius: 0.25rem;
        font-size: 28px;
        text-align: center;
        font-weight: bold;
    ">
        {text}
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)


# Display a message to confirm
with row3_col2:
    if queried_gene:
        styled_message(f"Gene: {queried_gene}", "success")
    elif queried_gene in ["", " "]:
        styled_message("Gene: -----", "warning")
    else:
        styled_message("GENE NOT FOUND", "error")
with row3_col3:
    if queried_tube_id:
        styled_message(f"Tube ID: {queried_tube_id}", "success")
    elif queried_tube_id in ["", " "]:
        styled_message("Tube ID: -----", "warning")
    else:
        styled_message("TUBE ID NOT FOUND", "error")

# Show button to add selected tube to selected spot
with row3_col4:
    if st.button("Confirm", type="primary"):
        if not gene_identifier or not st.session_state.selected_spot:
            st.error("Please enter an identifier and select a spot.")
        else:
            # --- Cross-Type Logging Validation ---
            # Check if a gene was found and if its type matches the selected log file.
            if queried_gene:  # Only run check if a gene was found
                is_wgk_log = gene_type == "WGK"
                is_batch_log = gene_type == "BATCH"
                gene_is_batch = str(queried_gene).strip().upper().startswith("BATCH")
                gene_is_wgk = str(queried_gene).strip().upper().startswith("WGK")

                if (is_wgk_log and gene_is_batch) or (is_batch_log and gene_is_wgk):
                    st.error(
                        f"Cross-logging error: Cannot add a '{'BATCH' if gene_is_batch else 'WGK'}' gene to a '{gene_type}' log."
                    )
                    st.stop()  # Halt execution to prevent incorrect logging
            # --- End Validation ---

            spot = st.session_state.selected_spot

            # Use queried results if available
            gene_val = (
                queried_gene
                if queried_gene
                else (gene_identifier if id_type == "Gene" else "")
            )
            tube_val = (
                queried_tube_id
                if queried_tube_id
                else (gene_identifier if id_type == "Tube ID" else "")
            )

            idx_to_update = current_box_df[current_box_df["Location"] == spot].index[0]
            current_box_df.at[idx_to_update, "Gene"] = gene_val
            current_box_df.at[idx_to_update, "Tube ID"] = tube_val
            current_box_df.at[idx_to_update, "Date"] = (
                queried_date if queried_date else ""
            )
            current_box_df.at[idx_to_update, "Amount"] = amount
            current_box_df.at[idx_to_update, "Notes"] = notes

            st.session_state.box_df = current_box_df.copy()
            st.session_state.selected_spot = None

            st.rerun()

# --- Editable table ---
st.markdown("### Current Box Table (Editable)")
edited_df = st.data_editor(
    st.session_state.box_df,
    width="stretch",
    hide_index=True,
    column_config={
        "Box name": st.column_config.TextColumn(disabled=True),
        "Location": st.column_config.TextColumn(disabled=True),
    },
)

# Update session state with any manual edits
st.session_state.box_df = edited_df

# --- Save to Excel ---
if st.button("Save to Log", type="primary"):
    try:

        # Clean identifier rows
        edited_df_cleaned = edited_df.copy()
        id_mask = edited_df_cleaned["Gene"].apply(is_nonempty) | edited_df_cleaned[
            "Tube ID"
        ].apply(is_nonempty)

        columns_to_clear = edited_df_cleaned.columns[2:]
        for idx, has_id in id_mask.items():
            if not has_id:
                edited_df_cleaned.loc[idx, columns_to_clear] = ""

        # Update session state with cleaned DataFrame so UI reflects changes immediately
        st.session_state.box_df = edited_df_cleaned.copy()

        # Step 2: Load full log and update box data
        full_df = pd.read_excel(filename, sheet_name="Log")
        full_df = full_df.fillna("")

        # Find the indices of the current box in full_df
        box_mask = full_df.iloc[:, 0] == box_name
        box_indices = full_df.index[box_mask]

        if len(box_indices) > 0:
            first_idx = box_indices.min()
            last_idx = box_indices.max()
            # Split into three parts
            before = (
                full_df.loc[: first_idx - 1]
                if first_idx > 0
                else pd.DataFrame(columns=full_df.columns)
            )
            after = (
                full_df.loc[last_idx + 1 :]
                if last_idx < full_df.index.max()
                else pd.DataFrame(columns=full_df.columns)
            )

            # Concatenate: before + edited box + after
            updated_df = pd.concat(
                [before, edited_df_cleaned, after], ignore_index=True
            )
        else:
            # Box not found, just append at the end (for new boxes)
            updated_df = pd.concat([full_df, edited_df_cleaned], ignore_index=True)

        # Step 3: Save and format
        save_and_format_log(filename, updated_df)

        st.success(f"Changes saved successfully to '{filename}'")
        # After saving locally, copy the updated file to the network drive
        copy_log_to_network(filename)

        st.rerun()

    except Exception as e:
        st.error(f"Failed to save changes: {e}")
